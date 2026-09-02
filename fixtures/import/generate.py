#!/usr/bin/env python3
"""Regenerate the *binary* import fixtures.

The delimited fixtures that are plain UTF-8 text are committed as text and edited
by hand. The ones below cannot be: a workbook is a zip or a BIFF stream, and the
two non-UTF-8 CSVs exist precisely because their bytes are not text an editor
would preserve. They are committed as binaries and this script is how they are
rebuilt, so a reviewer can see what is in them without a hex editor.

Run from the repository root:

    nix-shell -p python3Packages.openpyxl python3Packages.xlwt python3Packages.odfpy \
        --run "python3 fixtures/import/generate.py"

Every value here is synthetic. See fixtures/import/README.md for what each file
proves.
"""

from __future__ import annotations

import datetime
import pathlib

import openpyxl
import xlwt
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P

ROOT = pathlib.Path(__file__).resolve().parent
DELIMITED = ROOT / "delimited"
SPREADSHEET = ROOT / "spreadsheet"

# One statement, shared by every spreadsheet fixture, so a test that compares two
# backends is comparing the backends and not the data.
STATEMENT = [
    ["Date", "Description", "Amount", "Balance"],
    [datetime.date(2026, 1, 5), "GROCERY STORE", -54.20, 1200.00],
    [datetime.date(2026, 1, 6), "ATM WITHDRAWAL", -100.00, 1100.00],
    [datetime.date(2026, 1, 7), "EMPLOYER PAYROLL", 2500.00, 3600.00],
    [datetime.date(2026, 1, 9), "CORNER MARKET", -31.18, 3568.82],
]

# A currency number format on the Amount column. It is here to be IGNORED:
# calamine exposes no number formats at all, so the fixture proves we emit
# `-54.2` and never `($54.20)`.
CURRENCY = '"$"#,##0.00_);("$"#,##0.00)'


def simple_xlsx() -> None:
    """Header row plus date cells stored as serial numbers with a date format."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Statement"
    for row in STATEMENT:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = CURRENCY
    book.save(SPREADSHEET / "simple.xlsx")


def multi_sheet_xlsx() -> None:
    """Three sheets: one that is not a table, and two that are.

    `Cover` is a single populated cell, so sheet selection must walk past it.
    `Transactions` starts at C4, so the blank leading rows and columns must be
    trimmed before it looks like a table at all. `Summary` is a second genuine
    candidate, which is what makes `SheetChosen` owed to the user.
    """
    book = openpyxl.Workbook()
    cover = book.active
    cover.title = "Cover"
    cover["A1"] = "Acme Bank - statement export"

    transactions = book.create_sheet("Transactions")
    for offset, row in enumerate(STATEMENT):
        for column, value in enumerate(row):
            transactions.cell(row=4 + offset, column=3 + column, value=value)
    for offset in range(1, len(STATEMENT)):
        transactions.cell(row=4 + offset, column=3).number_format = "yyyy-mm-dd"

    summary = book.create_sheet("Summary")
    for row in [["Opening", 1254.20], ["Closing", 3568.82]]:
        summary.append(row)

    book.save(SPREADSHEET / "multi-sheet.xlsx")


def preamble_xlsx() -> None:
    """A floating title block above the header, which is what a real export ships.

    The shape is taken cell-for-cell from a real brokerage "All Activity" export:
    two empty rows, a one-cell title, an empty row, a second one-cell title,
    another empty row, and only then the column labels. Trimming the *blank*
    edges is not enough here — every one of those rows is inside the trimmed
    rectangle, so "the first populated row is the header" reads `All Activity
    Types` as the header and the entire statement as its body.

    The table below the preamble is the shared STATEMENT, so the conversion must
    produce a `Tabular` equal to `simple.xlsx`'s — differing only by
    `PreambleSkipped { lines: 4 }`.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "AllActivity"
    sheet["A3"] = "All Activity Types"
    sheet["A5"] = "Account Activity for All Accounts from Last 30 Days"

    first = 7
    for offset, row in enumerate(STATEMENT):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=first + offset, column=column, value=value)
    for offset in range(1, len(STATEMENT)):
        sheet.cell(row=first + offset, column=1).number_format = "yyyy-mm-dd"
        for column in (3, 4):
            sheet.cell(row=first + offset, column=column).number_format = CURRENCY

    book.save(SPREADSHEET / "preamble.xlsx")


def trailer_xlsx() -> None:
    """A disclaimer block BELOW the transactions, which is what a real export ships.

    The synthetic twin of `real-brokerage-preamble.xlsx`'s other end, and it
    carries three traps in one sheet on purpose:

    1. A **trailer** of blank rows and one-cell paragraphs under the last
       transaction. Left in place, `to_csv` renders each of them as a record of
       empty fields, and hledger abandons the *entire* file on the first one —
       so a correct rules file reports that the data will not parse.
    2. A **blank row inside** the transactions. Same failure, different place, so
       trimming only the end is not enough.
    3. The last transaction has an **empty final column**, so it reaches column
       three of four. A rule spelled "narrower than the header" trims it and the
       user silently loses a transaction; the rule has to be "too narrow to hold
       a date and an amount". This is the row the trim must stop at.

    The four surviving rows are the shared STATEMENT's, so the conversion must
    produce `simple.xlsx`'s table with its last Balance blanked, plus
    `TrailerSkipped {lines: 4}` and `BlankRowsDropped {count: 1}`.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Statement"

    # Row 4 is left empty on purpose: the blank row inside the transactions.
    layout = [(1, STATEMENT[0]), (2, STATEMENT[1]), (3, STATEMENT[2]), (5, STATEMENT[3])]
    # The last transaction, with its Balance cleared.
    layout.append((6, STATEMENT[4][:3] + [None]))
    for at, row in layout:
        for column, value in enumerate(row, start=1):
            if value is not None:
                sheet.cell(row=at, column=column, value=value)
    for at in (2, 3, 5, 6):
        sheet.cell(row=at, column=1).number_format = "yyyy-mm-dd"
        for column in (3, 4):
            sheet.cell(row=at, column=column).number_format = CURRENCY

    # Rows 7 and 9 are blank; 8 and 10 are the disclaimer paragraphs. The block
    # has to END on a populated row, or trimming the sheet's blank edges would
    # dispose of it before the rule under test ever sees it.
    sheet["A8"] = "*Balances shown are as of the statement date and may not reflect pending activity."
    sheet["A10"] = "Acme Bank is a member FDIC. Please retain this statement for your records."

    book.save(SPREADSHEET / "trailer.xlsx")


def single_column_xlsx() -> None:
    """A genuine one-column sheet, under a title of its own.

    The counter-example to the preamble rule. Every row here holds exactly one
    populated cell, so a rule spelled "a row with one cell in it is a title"
    discards the whole sheet — and a rule that instead looked for the first row
    matching the modal width would find it on line one and skip nothing, which is
    the answer we want. A one-wide table carries no signal either way, so this
    file must come back as `NoTable`: a sheet with no date and no amount is not a
    statement, and it must not be *reshaped* into one.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Balances"
    sheet["A1"] = "Daily Closing Balance"
    sheet["A3"] = "Balance"
    for offset, value in enumerate([1200.00, 1100.00, 3600.00, 3568.82, 3510.00]):
        sheet.cell(row=4 + offset, column=1, value=value)
    book.save(SPREADSHEET / "single-column.xlsx")


def no_table_xlsx() -> None:
    """A workbook that opens cleanly and contains no table.

    A cover note and an empty sheet. This is `ConvertError::NoTable` — a specific
    answer about a perfectly valid file, not a parse failure — and it exists so a
    test can tell the two apart.
    """
    book = openpyxl.Workbook()
    cover = book.active
    cover.title = "Cover"
    cover["A1"] = "Acme Bank"
    cover["A2"] = "Please retain for your records."
    book.create_sheet("Blank")
    book.save(SPREADSHEET / "no-table.xlsx")


def legacy_xls() -> None:
    """The BIFF8 path, which is a completely different reader inside calamine."""
    book = xlwt.Workbook()
    sheet = book.add_sheet("Statement")
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    money_style = xlwt.easyxf(num_format_str=CURRENCY)
    for at, row in enumerate(STATEMENT):
        for column, value in enumerate(row):
            if isinstance(value, datetime.date):
                sheet.write(at, column, value, date_style)
            elif isinstance(value, float):
                sheet.write(at, column, value, money_style)
            else:
                sheet.write(at, column, value)
    book.save(str(SPREADSHEET / "legacy.xls"))


def sheet_ods() -> None:
    """ODS, where a date is ISO 8601 text in an attribute and never a serial."""
    document = OpenDocumentSpreadsheet()
    table = Table(name="Statement")
    for row in STATEMENT:
        line = TableRow()
        for value in row:
            if isinstance(value, datetime.date):
                cell = TableCell(valuetype="date", datevalue=value.isoformat())
                cell.addElement(P(text=value.isoformat()))
            elif isinstance(value, float):
                cell = TableCell(valuetype="float", value=value)
                cell.addElement(P(text=f"{value:.2f}"))
            else:
                cell = TableCell(valuetype="string")
                cell.addElement(P(text=str(value)))
            line.addElement(cell)
        table.addElement(line)
    document.spreadsheet.addElement(table)
    document.save(str(SPREADSHEET / "sheet.ods"))


def latin1_csv() -> None:
    """Windows-1252, with every byte that separates it from ISO-8859-1.

    0x92, 0x93, 0x94 and 0x80 are a right single quote, a pair of double quotes
    and the euro sign in Windows-1252, and unassigned C1 control characters in
    ISO-8859-1. None of them is valid UTF-8, so chardetng is the only thing that
    can decode this file, and only one of its two plausible answers is right.
    """
    # Written as the characters, encoded to the bytes: U+2019 -> 0x92,
    # U+201C/U+201D -> 0x93/0x94, U+20AC -> 0x80, U+00C9 -> 0xC9.
    text = (
        "Date,Description,Amount\r\n"
        "2026-01-05,MCDONALD’S RESTAURANT,-8.40\r\n"
        "2026-01-06,“THE CORNER” DELICATESSEN,-12.25\r\n"
        "2026-01-07,CAFÉ RÉPUBLIQUE PARIS,-31.00\r\n"
        "2026-01-08,ACME GMBH INVOICE €50 FEE,-50.00\r\n"
    )
    encoded = text.encode("cp1252")
    for byte in (0x92, 0x93, 0x94, 0x80, 0xC9):
        assert byte in encoded, f"0x{byte:02X} must survive into the fixture"
    (DELIMITED / "latin1.csv").write_bytes(encoded)


def utf16le_bom_csv() -> None:
    """What Excel's "Unicode Text" export writes, saved under a .csv name.

    UTF-16LE with a byte-order mark and CRLF terminators. Handed to chardetng
    first this decodes as windows-1252 and every cell comes back with a NUL after
    every character; the BOM has to be believed before any detector is consulted.
    """
    text = (
        "Date,Description,Amount\r\n"
        "2026-01-05,CAFÉ RÉPUBLIQUE,-8.40\r\n"
        "2026-01-06,BÜCHER MÜNCHEN,-12.25\r\n"
        "2026-01-07,EMPLOYER PAYROLL,2500.00\r\n"
    )
    (DELIMITED / "utf16le-bom.csv").write_bytes(b"\xff\xfe" + text.encode("utf-16-le"))


# ---------------------------------------------------------------------------
# qb-journal/ — QuickBooks Online "Journal" report exports
# ---------------------------------------------------------------------------
#
# Every shape below was measured against a REAL QuickBooks Online Journal export
# (204 rows, 46 groups) with both openpyxl and this repo's own calamine. The data
# is scrubbed — no real company, payee, account or balance survives — but every
# structural property does, and the ones that are load-bearing are called out at
# each builder. See fixtures/import/qb-journal/README.md.

QB = ROOT / "qb-journal"

# The customized column set: what QuickBooks Online's report column picker
# produced for the motivating export. Column A has NO header name at all — that
# is where the transaction id and the "Total for {id}" text live.
QB_CUSTOM_HEADER = [
    "",
    "Transaction date",
    "Transaction type",
    "Num",
    "Name",
    "Description",
    "Distribution account number",
    "Account Name",
    "Debit",
    "Credit",
    "Item class",
    "Balance",
    "Customer full name",
    "Vendor",
]

# The stock column set, with QuickBooks' other spellings for two of the columns.
# Four columns fewer AND two different names, so a detector or a column mapper
# keyed on an exact header list or a column count fails on this and only this.
QB_DEFAULT_HEADER = [
    "",
    "Transaction date",
    "Transaction type",
    "Num",
    "Name",
    "Memo/Description",
    "Account",
    "Debit",
    "Credit",
]

QB_NUMBER = "#,##0.00"
QB_TOTAL_NUMBER = '"$"#,##0.00'


# openpyxl drops a cell whose value is "" rather than writing an empty one, so
# every blank text cell goes in under this sentinel and `_qb_patch_formulas`
# rewrites it to a genuinely empty inline string on the way out.
QB_BLANK = "@@BLANK@@"


def _qb_posting(row, width):
    """Pad one posting row out to `width`, spelling a blank cell empty-but-present.

    Not cosmetic. In the real export every unused text cell on a posting row is
    an empty *string* from the shared-string table, which calamine hands over as
    `Data::String("")` and not `Data::Empty`. So "this column has nothing in it"
    has to mean "holds nothing printable" everywhere it is asked — a reader that
    tests for `Data::Empty` alone reports a Vendor of `Some("")` on every row
    that has no vendor, and carries an empty tag into the written journal.

    The marker column is the exception and stays genuinely absent, because that
    is what the real export writes there: `Data::Empty` in column A on a posting
    row, empty *strings* in the unused columns to its right.
    """
    padded = [QB_BLANK if value == "" else value for value in row]
    padded += [QB_BLANK] * (width - len(padded))
    return [None] + padded[1:]


def _qb_sheet(book, header, groups, *, title="Northwind Widgets LLC", total=None,
              footer=" Wednesday, September 02, 2026 10:31 AM GMT-06:00"):
    """Lay a QuickBooks Journal report out on `book`'s active sheet.

    `groups` is a list of (marker_text, posting_rows, total_row) where
    `total_row` is (total_text, debit_cached, credit_cached). A cached value
    given as a *string* is written into the XML verbatim, which is how a fixture
    carries the seventeen digits Excel stores (`70120.850000000006`) rather than
    the shortest form Python would print; "#REF!" additionally becomes an error
    cell. Returns the {cell_ref: cached} map the formula patch below needs.
    """
    sheet = book.active
    sheet.title = "Sheet1"
    width = len(header)

    # Rows 1-3 are a merged title band and row 4 is blank: the ordinary report
    # preamble, and the reason the header is not row 1.
    for offset, line in enumerate((title, "Journal", "All Dates"), start=1):
        sheet.cell(row=offset, column=1, value=line)
        sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=width - 1)
    for column, label in enumerate(header, start=1):
        if label:
            sheet.cell(row=5, column=column, value=label)

    # The two amount columns come from the header, not from a constant: the stock
    # column set puts Debit and Credit one place to the left of the customized one.
    debit_at = header.index("Debit") + 1
    credit_at = header.index("Credit") + 1

    cached = {}
    at = 6
    for marker, postings, (total_text, debit, credit) in groups:
        first = at + 1
        sheet.cell(row=at, column=1, value=marker)
        at += 1
        for posting in postings:
            for column, value in enumerate(_qb_posting(posting, width), start=1):
                cell = sheet.cell(row=at, column=column, value=value)
                if isinstance(value, float) or (isinstance(value, int) and column > 8):
                    cell.number_format = QB_NUMBER
            at += 1
        last = at - 1
        sheet.cell(row=at, column=1, value=total_text)
        # The total row's two numbers are FORMULAS in a real export, and the
        # number that reaches us is the formula's *cached* value. Written that
        # way here (see `_qb_patch_formulas`) so a reader that ignores `<v>` when
        # `<f>` is present, or that tries to evaluate the formula, fails.
        for column, value in ((debit_at, debit), (credit_at, credit)):
            if value is None:
                continue
            ref = f"{chr(ord('A') + column - 1)}{at}"
            span = "+".join(f"{ref[0]}{n}" for n in range(first, last + 1))
            formula = "=#REF!+#REF!" if value == "#REF!" else f"={span}"
            cell = sheet.cell(row=at, column=column, value=formula)
            cell.number_format = QB_TOTAL_NUMBER
            cached[ref] = value
        at += 1

    if total is not None:
        sheet.cell(row=at, column=1, value="TOTAL")
        for column, value in ((debit_at, total), (credit_at, total)):
            cell = sheet.cell(row=at, column=column, value=value)
            cell.number_format = QB_TOTAL_NUMBER
        at += 1
    if footer is not None:
        at += 3
        sheet.cell(row=at, column=1, value=footer)
        sheet.merge_cells(start_row=at, start_column=1, end_row=at, end_column=width - 1)
    return cached


def _qb_patch_formulas(path, cached):
    """Give every formula cell in `path` the cached value openpyxl will not write.

    openpyxl writes `<f>…</f>` and stops, which is a workbook no spreadsheet has
    ever produced: Excel always stores the last computed value alongside the
    formula, and that stored value is the only thing calamine (or any reader that
    does not evaluate formulas) can see. Rewriting the XML is the only way to
    commit a fixture shaped like the real file.

    A cached "#REF!" additionally needs `t="e"`, which is how a formula error is
    spelled on the wire and what makes calamine yield `Data::Error(Ref)` rather
    than the string.
    """
    import re
    import shutil
    import zipfile

    source = pathlib.Path(str(path) + ".tmp")
    shutil.move(path, source)
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(
        path, "w", zipfile.ZIP_DEFLATED
    ) as patched:
        for item in original.infolist():
            data = original.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                for ref, value in cached.items():
                    # openpyxl emits `<f>…</f><v />` — a formula with the cached
                    # value left empty, which is a workbook Excel never writes.
                    pattern = r'<c r="%s"([^>]*)>(<f>[^<]*</f>)(<v ?/>|<v>[^<]*</v>)?</c>' % ref
                    digits = value if isinstance(value, str) else repr(value)
                    attrs = r'\1 t="e"' if value == "#REF!" else r"\1"
                    text, count = re.subn(
                        pattern, r'<c r="%s"%s>\2<v>%s</v></c>' % (ref, attrs, digits), text
                    )
                    assert count == 1, f"{path.name}: {ref} not patched ({count} matches)"
                text = text.replace(f"<t>{QB_BLANK}</t>", "<t></t>")
                assert QB_BLANK not in text, f"{path.name}: blank sentinel survived"
                data = text.encode("utf-8")
            patched.writestr(item, data)
    source.unlink()


def _qb_write(name, header, groups, **kwargs):
    book = openpyxl.Workbook()
    cached = _qb_sheet(book, header, groups, **kwargs)
    path = QB / name
    book.save(path)
    _qb_patch_formulas(path, cached)


# Two groups whose shapes cover 45 of the real export's 46: a Deposit that funds
# a bank account out of equity, and an Expense charged to a credit card. Written
# once so `simple.xlsx` and `default-columns.xlsx` differ only in their columns.
QB_DEPOSIT = [
    ["", "01/17/2026", "Deposit", "", "Ridgeline Partners, LP", "Deposit 0000065913", "",
     "Riverbank BUSINESS CHECKING (0002)", 74999.71, None, "", 74999.71, "Ridgeline Partners, LP", ""],
    ["", "01/17/2026", "Deposit", "", "Ridgeline Partners, LP", "Deposit 0000065913", "3000",
     "3000 Member Equity", None, 74999.71, "", 149999.42, "Ridgeline Partners, LP", ""],
]
QB_EXPENSE = [
    ["", "01/05/2026", "Expense", "", "Grasshopper Cloud", "GRASSHOPPER-308*5498114", "2005",
     "2005 Northbank Credit Card", None, 79.99, "", 79.99, "", "Grasshopper Cloud"],
    ["", "01/05/2026", "Expense", "", "Grasshopper Cloud", "GRASSHOPPER-308*5498114", "6001",
     "6000 Sales and Marketing:6001 Sales & Marketing Tools", 79.99, None, "", 159.98, "",
     "Grasshopper Cloud"],
]


def _qb_narrow(rows):
    """The customized 14-column rows as the stock 9-column report would write them.

    The four added columns come off the right-hand end, and `Distribution account
    number` comes out of the middle — so this is not a truncation, and a fixture
    built by slicing would hand the parser an account name where its header says
    Debit.
    """
    return [row[:6] + row[7:10] for row in rows]


def qb_simple_xlsx() -> None:
    """The baseline: the title band, the customized columns, two ordinary groups.

    Both groups are the two-posting shape that is 44 of the real export's 46. The
    Deposit is the sign check that needs no account-type knowledge — the bank
    account is debited and equity is credited, so `debit if debit else -credit`
    has to yield +74999.71 and -74999.71 — and the Expense is the same check from
    the liability side.
    """
    _qb_write(
        "simple.xlsx",
        QB_CUSTOM_HEADER,
        [
            ("441", QB_DEPOSIT, ("Total for 441", 74999.71, 74999.71)),
            ("33", QB_EXPENSE, ("Total for 33", "79.989999999999995", "79.989999999999995")),
        ],
        total=75079.700000000001,
    )


def qb_default_columns_xlsx() -> None:
    """The same two groups under the stock column set, and two other spellings.

    `Item class`, `Balance`, `Customer full name` and `Vendor` are gone, and the
    two that remain are labelled `Memo/Description` and `Account` rather than
    `Description` and `Account Name`. Parsed, it must yield the same accounts,
    dates and signed amounts `simple.xlsx` does — so a detector keyed on the
    exact 14 labels, or on a column count, or a mapper that only knows one
    spelling of the memo and account columns, fails here and nowhere else.
    """
    _qb_write(
        "default-columns.xlsx",
        QB_DEFAULT_HEADER,
        [
            ("441", _qb_narrow(QB_DEPOSIT), ("Total for 441", 74999.71, 74999.71)),
            ("33", _qb_narrow(QB_EXPENSE), ("Total for 33", "79.989999999999995", "79.989999999999995")),
        ],
        total=75079.700000000001,
    )


# The manual Journal Entry: ten postings, `Num` populated on every one of them,
# a sub-account whose name contains a colon, and a Description that CHANGES from
# posting to posting. All four are real properties of the export's group 612.
QB_JOURNAL_ENTRY = [
    ["", "01/01/2026", "Journal Entry", "2", "", "Opening Balance Entry", "3000",
     "3000 Member Equity", None, 70000.0, "", 70000.0, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Opening Balance Entry", "3900",
     "3900 Retained Earnings", 35131.01, None, "", 34868.99, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Bulk Buy - Soda Machine", "1520",
     "1520 Computer & Office Equipment", 49.99, None, "", 34918.979999999996, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Riverbend - LCD Monitor", "1520",
     "1520 Computer & Office Equipment", 338.98, None, "", 35257.96, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Monitor Arms", "1520",
     "1520 Computer & Office Equipment", 956.0, None, "", 36213.96, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Riverbend", "1520",
     "1520 Computer & Office Equipment", 17.98, None, "", 36231.94, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "", "Opening Balance Entry", "",
     "Opening Balance Equity", 33571.04, None, "", 2660.9000000000015, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "",
     "Opening Balance Entry - Accumulated Depreciation catchup", "1521",
     "1520 Computer & Office Equipment:1521 Computer & Equipment - Accum Depr", None, 55.85, "",
     2605.0500000000015, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "",
     "Opening Balance Entry - Accumulated Depreciation catchup", "3900",
     "3900 Retained Earnings", 55.85, None, "", 2549.2000000000016, "", ""],
    ["", "01/01/2026", "Journal Entry", "2", "",
     "Opening Balance Entry - write off 2015 reconciling items", "3900",
     "3900 Retained Earnings", None, 65.0, "", 2614.2000000000016, "", ""],
]


def qb_many_postings_xlsx() -> None:
    """Ten postings under one marker, and the two traps that come with them.

    The Description differs on six of the ten, so a parser that reads the memo
    once for the group and hangs it off the transaction loses eight distinct
    memos and keeps a wrong one. And the total row's cached value is written as
    the seventeen digits Excel actually stored — `70120.850000000006` — so a
    reader that compares the total against its own sum in `f64`, or that reads
    the stored digits as text, disagrees with the postings by 6e-12 and refuses
    a perfectly good file. Shortest-round-trip printing is what recovers 70120.85
    exactly, and that is the only thing that does.
    """
    _qb_write(
        "many-postings.xlsx",
        QB_CUSTOM_HEADER,
        [("612", QB_JOURNAL_ENTRY, ("Total for 612", "70120.850000000006", "70120.850000000006"))],
        total=70120.850000000006,
    )


# The four-posting Bill the real export's truncation left stranded: one payable
# credited, three legal-fee debits under distinct invoice memos.
QB_BILL = [
    ["", "01/20/2026", "Bill", "", "Harbor & Vance LLP", "", "2000", "2000 Accounts Payable",
     None, 533.94, "", 533.94, "", "Harbor & Vance LLP"],
    ["", "01/20/2026", "Bill", "", "Harbor & Vance LLP", "Invoice #24688 8/4/25", "6310",
     "6300 Professional Fees:6310 Legal Fees", 341.5, None, "", 875.44, "", "Harbor & Vance LLP"],
    ["", "01/20/2026", "Bill", "", "Harbor & Vance LLP", "Invoice #24797 9/7/25", "6310",
     "6300 Professional Fees:6310 Legal Fees", 105.0, None, "", 980.44, "", "Harbor & Vance LLP"],
    ["", "01/20/2026", "Bill", "", "Harbor & Vance LLP", "Invoice #24499 6/6/25", "6310",
     "6300 Professional Fees:6310 Legal Fees", 87.44, None, "", 1067.88, "", "Harbor & Vance LLP"],
]


def qb_truncated_tail_xlsx() -> None:
    """The real export's own damage, reproduced: a group closed by another's total.

    The sample this corpus was built from had been hand-truncated in Excel, and
    the tell is not the `#REF!` — it is that the surviving `Total for 11024`
    closes a group whose marker says `6`. The postings under it balance perfectly
    (533.94 both ways), so every arithmetic check passes and only the id says the
    file lost rows. A parser that pairs marker to total by POSITION rather than by
    id imports this as transaction 6 and is wrong about it in silence.
    """
    _qb_write(
        "truncated-tail.xlsx",
        QB_CUSTOM_HEADER,
        [
            ("441", QB_DEPOSIT, ("Total for 441", 74999.71, 74999.71)),
            ("6", QB_BILL, ("Total for 11024", "#REF!", "#REF!")),
        ],
        total=65510189.670000099,
    )


def qb_malformed_total_xlsx() -> None:
    """A `#REF!` total on a group whose id is intact.

    The other half of `truncated-tail.xlsx`: same broken cells, matching ids, so
    this is the fixture that reaches the total-cell check at all. `#REF!` reaches
    us as `Data::Error(Ref)`, and `convert::spreadsheet` renders exactly that as
    an empty string — which here would read as a total of zero against postings
    of 533.94, or as no total at all. It has to be a named refusal.
    """
    _qb_write(
        "malformed-total.xlsx",
        QB_CUSTOM_HEADER,
        [("6", QB_BILL, ("Total for 6", "#REF!", "#REF!"))],
        total=533.94,
    )


def qb_mismatched_total_xlsx() -> None:
    """A total row that disagrees with the postings it claims to add up.

    Constructed, not observed: in an untouched export the total is a formula over
    the very rows above it and cannot disagree. It CAN disagree the moment a
    human edits an amount in a spreadsheet that does not recalculate on open,
    which leaves a stale cached value — exactly the shape written here (the
    postings sum to 533.94 and the stored total says 500.00). Nothing else in the
    file is wrong, so this is the only fixture that fails a parser which trusts
    its own sum and never reads the total row at all.
    """
    _qb_write(
        "mismatched-total.xlsx",
        QB_CUSTOM_HEADER,
        [("6", QB_BILL, ("Total for 6", 500.0, 500.0))],
        total=500.0,
    )


def qb_orphan_total_xlsx() -> None:
    """A `Total for 99` closing nothing, before any marker row has opened.

    The truncation failure from the other end — rows deleted from the TOP of a
    group. A parser that treats "Total for" as merely "flush whatever postings I
    am holding" emits a transaction with no id here, or silently emits nothing.
    """
    book = openpyxl.Workbook()
    cached = _qb_sheet(
        book,
        QB_CUSTOM_HEADER,
        [("441", QB_DEPOSIT, ("Total for 441", 74999.71, 74999.71))],
        total=74999.71,
    )
    sheet = book.active
    # Splice the orphan in directly above the legitimate group's marker.
    sheet.insert_rows(6)
    sheet.cell(row=6, column=1, value="Total for 99")
    sheet.cell(row=6, column=9, value=1234.56).number_format = QB_TOTAL_NUMBER
    sheet.cell(row=6, column=10, value=1234.56).number_format = QB_TOTAL_NUMBER
    path = QB / "orphan-total.xlsx"
    book.save(path)
    # `insert_rows` moved every patched cell down one row.
    _qb_patch_formulas(path, {f"{ref[0]}{int(ref[1:]) + 1}": v for ref, v in cached.items()})


def qb_overlap_xlsx() -> None:
    """A WIDER re-download: one group already imported from `simple.xlsx`, one new.

    Built for Phase B's server-side write pipeline, not Phase A's parser: the
    "re-downloading is safe" property the plan documents rests on a single
    export mixing ids the journal already holds with ids it does not, in ONE
    commit. Group `441` is `QB_DEPOSIT` byte-for-byte — the same transaction
    `simple.xlsx` also carries under that id, so importing `simple.xlsx` and
    then this file must classify it `Unchanged` and write nothing for it — and
    group `6` is `QB_BILL`, an id neither `simple.xlsx` nor `default-columns.xlsx`
    ever uses, so it is the one row a commit of this file actually writes.
    """
    _qb_write(
        "overlap.xlsx",
        QB_CUSTOM_HEADER,
        [
            ("441", QB_DEPOSIT, ("Total for 441", 74999.71, 74999.71)),
            ("6", QB_BILL, ("Total for 6", 533.94, 533.94)),
        ],
        total=75533.65,
    )


def qb_near_miss_xlsx() -> None:
    """NOT a QuickBooks Journal, and the reason detection cannot stop at the header.

    An ordinary bank export that happens to carry `Account Name`, `Debit` AND
    `Credit` labels — the exact triple the QuickBooks header is recognised by —
    and closes with a `Total` summary row carrying two numbers. Everything a
    detector keyed on column names alone would want is here. What is absent is
    the only thing that actually identifies the format: a bare-id marker row, and
    a closing row that says `Total for` some id. Detection must answer no.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Activity"
    rows = [
        ["Date", "Description", "Account Name", "Debit", "Credit", "Balance"],
        ["2026-01-05", "GROCERY STORE", "Checking (0002)", None, 54.20, 1200.00],
        ["2026-01-06", "ATM WITHDRAWAL", "Checking (0002)", None, 100.00, 1100.00],
        ["2026-01-07", "EMPLOYER PAYROLL", "Checking (0002)", 2500.00, None, 3600.00],
        ["Total", None, None, 2500.00, 154.20, None],
    ]
    for row in rows:
        sheet.append(row)
    book.save(QB / "near-miss.xlsx")


def qb_report_xlsx() -> None:
    """The whole export's shape at full size: 45 groups, 100 posting rows.

    The corpus-level fixture. Group sizes and the type mix are the real file's
    once its one damaged group is set aside — 43 two-posting groups, one Bill of
    four and one Journal Entry of ten — so the round-trip property has something
    to be true of: the number of transactions parsed equals the number of
    `Total for ` rows anyone can count in a spreadsheet, and every one of them
    balances. Dates run past the 12th, which is what makes `01/17/2026`
    month-first rather than a coin toss.
    """
    types = [
        ("Deposit", "Riverbank BUSINESS CHECKING (0002)", "3000 Member Equity", True),
        ("Expense", "2005 Northbank Credit Card", "6100 G&A:6130 Bank Service Charges", False),
        ("Transfer", "Riverbank BUSINESS CHECKING (0002)", "2005 Northbank Credit Card", False),
        ("Credit Card Expense", "2005 Northbank Credit Card",
         "6600 Facilities & Ops:6620 Software and Cloud Services", False),
        ("Bill", "2000 Accounts Payable", "6300 Professional Fees:6310 Legal Fees", False),
    ]
    groups = []
    for index in range(43):
        kind, first, second, first_debit = types[index % len(types)]
        amount = round(25.0 + index * 137.53, 2)
        day = 1 + (index % 20)
        date = f"01/{day:02d}/2026"
        # The running Balance column, float noise and all. It is signed by each
        # account's normal side in the real report, which is not knowable from
        # anything in the file — so it is written here as nonsense on purpose and
        # nothing may ever read it.
        rows = [
            ["", date, kind, "", f"Payee {index:02d}", f"Memo {index:02d}", "",
             first, amount if first_debit else None, None if first_debit else amount, "",
             amount + 1e-11, f"Payee {index:02d}" if kind == "Deposit" else "",
             "" if kind == "Deposit" else f"Payee {index:02d}"],
            ["", date, kind, "", f"Payee {index:02d}", f"Memo {index:02d}", second.split(" ")[0],
             second, None if first_debit else amount, amount if first_debit else None, "",
             amount * 2, f"Payee {index:02d}" if kind == "Deposit" else "",
             "" if kind == "Deposit" else f"Payee {index:02d}"],
        ]
        groups.append((str(100 + index), rows, (f"Total for {100 + index}", amount, amount)))
    groups.append(("612", QB_JOURNAL_ENTRY,
                   ("Total for 612", "70120.850000000006", "70120.850000000006")))
    groups.append(("6", QB_BILL, ("Total for 6", 533.94, 533.94)))
    _qb_write("report.xlsx", QB_CUSTOM_HEADER, groups, total=65510189.670000099)


def main() -> None:
    SPREADSHEET.mkdir(parents=True, exist_ok=True)
    DELIMITED.mkdir(parents=True, exist_ok=True)
    QB.mkdir(parents=True, exist_ok=True)
    for build in (
        simple_xlsx,
        multi_sheet_xlsx,
        preamble_xlsx,
        trailer_xlsx,
        single_column_xlsx,
        no_table_xlsx,
        legacy_xls,
        sheet_ods,
        latin1_csv,
        utf16le_bom_csv,
        qb_simple_xlsx,
        qb_default_columns_xlsx,
        qb_many_postings_xlsx,
        qb_truncated_tail_xlsx,
        qb_malformed_total_xlsx,
        qb_mismatched_total_xlsx,
        qb_orphan_total_xlsx,
        qb_overlap_xlsx,
        qb_near_miss_xlsx,
        qb_report_xlsx,
    ):
        build()
        print(f"wrote {build.__name__}")


if __name__ == "__main__":
    main()
